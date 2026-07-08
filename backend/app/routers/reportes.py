import io
from typing import Optional
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

router = APIRouter(prefix="/api/reportes", tags=["reportes"])


class ReporteDatosRequest(BaseModel):
    user_id: str
    cotizacion_ids: list[str] = []
    proyecto_id: Optional[str] = None
    titulo: str = "Reporte de Cotización"
    secciones: list[str] = ["resumen", "items", "comparativa", "proveedores", "plazos"]


@router.post("/datos")
async def datos_reporte(req: ReporteDatosRequest):
    """Recopila todos los datos necesarios para generar el reporte."""
    from app.services.supabase import get_supabase
    sb = get_supabase()

    datos: dict = {
        "titulo": req.titulo,
        "secciones": req.secciones,
        "fecha": __import__("datetime").date.today().isoformat(),
        "items": [],
        "proveedores_detalle": {},
        "proyecto": None,
    }

    # ── Datos de proyecto ──────────────────────────────────────────────────────
    if req.proyecto_id:
        proy = sb.table("proyectos").select("*").eq("id", req.proyecto_id).single().execute()
        if proy.data:
            datos["proyecto"] = proy.data
            items_proy = sb.table("items_proyecto").select("*, proveedores(*)").eq("proyecto_id", req.proyecto_id).order("orden").execute()
            for it in items_proy.data:
                cotizaciones = sb.table("cotizaciones_proyecto").select("*").eq("item_proyecto_id", it["id"]).order("precio_unitario").execute()
                datos["items"].append({
                    "item": it.get("item"),
                    "descripcion": it.get("descripcion"),
                    "cantidad": it.get("cantidad", 1),
                    "unidad": it.get("unidad", "unidad"),
                    "precio_unitario": it.get("precio_unitario"),
                    "precio_total": it.get("precio_total"),
                    "plazo_entrega_dias": it.get("plazo_entrega_dias"),
                    "proveedor_seleccionado": (it.get("proveedores") or {}).get("nombre"),
                    "proveedor_id": it.get("proveedor_seleccionado_id"),
                    "cotizaciones": cotizaciones.data,
                })

    # ── Datos de cotizaciones individuales ────────────────────────────────────
    for cot_id in req.cotizacion_ids:
        cot = sb.table("cotizaciones").select("*").eq("id", cot_id).single().execute()
        resultados = sb.table("resultados").select("*").eq("cotizacion_id", cot_id).order("precio").execute()
        if cot.data:
            datos["items"].append({
                "item": cot.data.get("nombre_identificado") or cot.data.get("descripcion"),
                "descripcion": cot.data.get("descripcion"),
                "cantidad": 1,
                "unidad": "unidad",
                "precio_unitario": resultados.data[0]["precio"] if resultados.data else None,
                "precio_total": resultados.data[0]["precio"] if resultados.data else None,
                "plazo_entrega_dias": None,
                "proveedor_seleccionado": resultados.data[0]["proveedor_nombre"] if resultados.data else None,
                "cotizaciones": [{"proveedor_nombre": r["proveedor_nombre"], "precio_unitario": r["precio"], "fuente": r["fuente"]} for r in resultados.data],
            })

    # ── Info de proveedores involucrados ──────────────────────────────────────
    proveedores_ids = set()
    for it in datos["items"]:
        if it.get("proveedor_id"):
            proveedores_ids.add(it["proveedor_id"])

    for pid in proveedores_ids:
        prov = sb.table("proveedores").select("*").eq("id", pid).single().execute()
        if prov.data:
            ratings = sb.table("supplier_ratings").select("estrellas").eq("proveedor_id", pid).execute()
            ocs = sb.table("ordenes_compra").select("id, created_at").eq("proveedor_nombre", prov.data.get("nombre", "")).execute()
            prov.data["total_ratings"] = len(ratings.data)
            prov.data["rating_promedio"] = round(sum(r["estrellas"] for r in ratings.data) / len(ratings.data), 1) if ratings.data else None
            prov.data["total_ocs_historial"] = len(ocs.data)
            prov.data["es_nuevo"] = len(ocs.data) == 0
            datos["proveedores_detalle"][pid] = prov.data

    # ── Resumen ───────────────────────────────────────────────────────────────
    items_con_precio = [it for it in datos["items"] if it.get("precio_total")]
    datos["resumen"] = {
        "total_items": len(datos["items"]),
        "items_cotizados": len(items_con_precio),
        "monto_total": sum(float(it.get("precio_total") or 0) for it in datos["items"]),
        "proveedores_evaluados": sum(len(it.get("cotizaciones") or []) for it in datos["items"]),
        "max_plazo_dias": max((it.get("plazo_entrega_dias") or 0 for it in datos["items"]), default=0),
    }

    return datos


class ExcelRequest(BaseModel):
    user_id: str
    cotizacion_ids: list[str] = []
    proyecto_id: Optional[str] = None
    titulo: str = "Reporte de Cotización"


@router.post("/exportar-excel")
async def exportar_excel(req: ExcelRequest):
    """Genera Excel del reporte."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=503, detail="openpyxl no instalado")

    # Reusa la lógica de datos
    datos = await datos_reporte(ReporteDatosRequest(
        user_id=req.user_id,
        cotizacion_ids=req.cotizacion_ids,
        proyecto_id=req.proyecto_id,
        titulo=req.titulo,
    ))

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1E1E3A")

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # ── Hoja resumen ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.append(["Reporte", datos["titulo"]])
    ws1.append(["Fecha", datos["fecha"]])
    ws1.append([])
    ws1.append(["Total ítems", datos["resumen"]["total_items"]])
    ws1.append(["Ítems cotizados", datos["resumen"]["items_cotizados"]])
    ws1.append(["Monto total", datos["resumen"]["monto_total"]])
    ws1.append(["Proveedores evaluados", datos["resumen"]["proveedores_evaluados"]])

    # ── Hoja ítems ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle Ítems")
    ws2.append(["N°", "Ítem", "Descripción", "Cantidad", "Unidad", "Proveedor", "Precio Unit.", "Precio Total", "Plazo (días)"])
    style_header(ws2)
    for i, it in enumerate(datos["items"], 1):
        ws2.append([i, it.get("item"), it.get("descripcion"), it.get("cantidad"), it.get("unidad"), it.get("proveedor_seleccionado"), it.get("precio_unitario"), it.get("precio_total"), it.get("plazo_entrega_dias")])

    # ── Hoja comparativa ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Comparativa Proveedores")
    ws3.append(["Ítem", "Proveedor", "Precio Unitario", "Fuente"])
    style_header(ws3)
    for it in datos["items"]:
        for cot in (it.get("cotizaciones") or []):
            ws3.append([it.get("item"), cot.get("proveedor_nombre"), cot.get("precio_unitario"), cot.get("fuente")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_claria.xlsx"},
    )

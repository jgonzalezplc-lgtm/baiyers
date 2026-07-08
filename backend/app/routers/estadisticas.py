from fastapi import APIRouter
from fastapi.responses import StreamingResponse
import io

router = APIRouter(prefix="/api/estadisticas", tags=["estadisticas"])


@router.get("/resumen")
async def resumen(user_id: str):
    from app.services.supabase import get_supabase
    from datetime import date, timedelta
    sb = get_supabase()

    hoy = date.today()
    mes_inicio = hoy.replace(day=1).isoformat()
    mes_anterior_fin = (hoy.replace(day=1) - timedelta(days=1)).isoformat()
    mes_anterior_inicio = date(hoy.year if hoy.month > 1 else hoy.year - 1, (hoy.month - 2) % 12 + 1, 1).isoformat()

    # OCs este mes
    ocs_mes = sb.table("ordenes_compra").select("precio_total").eq("user_id", user_id).gte("created_at", mes_inicio).execute()
    total_ocs_mes = sum(float(o.get("precio_total") or 0) for o in ocs_mes.data)
    cant_ocs_mes = len(ocs_mes.data)

    # OCs mes anterior
    ocs_ant = sb.table("ordenes_compra").select("precio_total").eq("user_id", user_id).gte("created_at", mes_anterior_inicio).lte("created_at", mes_anterior_fin).execute()
    total_ocs_ant = sum(float(o.get("precio_total") or 0) for o in ocs_ant.data)

    # Facturas pendientes
    facturas_pend = sb.table("facturas").select("monto_total").eq("user_id", user_id).neq("estado", "pagada").execute()
    total_pendiente = sum(float(f.get("monto_total") or 0) for f in facturas_pend.data)

    # Facturas pagadas este mes
    facturas_pag = sb.table("facturas").select("monto_total").eq("user_id", user_id).eq("estado", "pagada").gte("fecha_pago", mes_inicio).execute()
    total_pagado = sum(float(f.get("monto_total") or 0) for f in facturas_pag.data)

    # Proyección próximo mes (recurrencias activas)
    recs = sb.table("recurrencias").select("monto_maximo").eq("user_id", user_id).eq("activa", True).execute()
    proyeccion = sum(float(r.get("monto_maximo") or 0) for r in recs.data)

    # Variación %
    variacion = 0.0
    if total_ocs_ant > 0:
        variacion = ((total_ocs_mes - total_ocs_ant) / total_ocs_ant) * 100

    return {
        "total_ocs_mes": total_ocs_mes,
        "cant_ocs_mes": cant_ocs_mes,
        "total_pendiente_pago": total_pendiente,
        "total_pagado_mes": total_pagado,
        "proyeccion_proximo_mes": proyeccion,
        "variacion_vs_anterior_pct": round(variacion, 1),
    }


@router.get("/gastos-mensuales")
async def gastos_mensuales(user_id: str):
    from app.services.supabase import get_supabase
    from datetime import date, timedelta
    sb = get_supabase()

    meses = []
    hoy = date.today()
    for i in range(11, -1, -1):
        mes = hoy.month - i
        año = hoy.year + (mes - 1) // 12
        mes = ((mes - 1) % 12) + 1
        meses.append((año, mes))

    resultado = []
    for año, mes in meses:
        inicio = f"{año}-{mes:02d}-01"
        import calendar
        ultimo_dia = calendar.monthrange(año, mes)[1]
        fin = f"{año}-{mes:02d}-{ultimo_dia}"

        ocs = sb.table("ordenes_compra").select("precio_total").eq("user_id", user_id).gte("created_at", inicio).lte("created_at", fin + "T23:59:59").execute()
        total = sum(float(o.get("precio_total") or 0) for o in ocs.data)

        resultado.append({
            "mes": f"{año}-{mes:02d}",
            "label": f"{['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'][mes-1]} {str(año)[2:]}",
            "total": total,
            "cantidad": len(ocs.data),
        })

    return resultado


@router.get("/por-categoria")
async def por_categoria(user_id: str):
    from app.services.supabase import get_supabase
    sb = get_supabase()

    ocs = sb.table("ordenes_compra").select("nombre_item, precio_total").eq("user_id", user_id).execute()

    categorias: dict = {}
    for oc in ocs.data:
        nombre = oc.get("nombre_item") or "Sin categoría"
        # Usar las primeras 2 palabras como categoría
        palabras = nombre.strip().split()
        cat = " ".join(palabras[:2]) if palabras else "Otro"
        monto = float(oc.get("precio_total") or 0)
        categorias[cat] = categorias.get(cat, 0) + monto

    return [
        {"categoria": k, "total": v}
        for k, v in sorted(categorias.items(), key=lambda x: x[1], reverse=True)[:10]
    ]


@router.get("/top-proveedores")
async def top_proveedores(user_id: str):
    from app.services.supabase import get_supabase
    sb = get_supabase()

    ocs = sb.table("ordenes_compra").select("proveedor_nombre, precio_total").eq("user_id", user_id).execute()

    provs: dict = {}
    for oc in ocs.data:
        nombre = oc.get("proveedor_nombre") or "Desconocido"
        monto = float(oc.get("precio_total") or 0)
        provs[nombre] = provs.get(nombre, 0) + monto

    return [
        {"proveedor": k, "total": v}
        for k, v in sorted(provs.items(), key=lambda x: x[1], reverse=True)[:5]
    ]


@router.get("/liquidez")
async def liquidez(user_id: str):
    """Facturas pendientes agrupadas por semana — próximos 90 días."""
    from app.services.supabase import get_supabase
    from datetime import date, timedelta
    sb = get_supabase()

    hoy = date.today()
    limite = (hoy + timedelta(days=90)).isoformat()

    facturas = sb.table("facturas").select("proveedor_nombre, monto_total, fecha_vencimiento").eq("user_id", user_id).neq("estado", "pagada").lte("fecha_vencimiento", limite).gte("fecha_vencimiento", hoy.isoformat()).execute()

    # Agrupar por semana
    semanas: dict = {}
    for f in facturas.data:
        if not f.get("fecha_vencimiento"):
            continue
        venc = date.fromisoformat(str(f["fecha_vencimiento"]))
        # Inicio de semana (lunes)
        lunes = venc - timedelta(days=venc.weekday())
        key = lunes.isoformat()
        if key not in semanas:
            semanas[key] = {"semana": key, "total": 0, "facturas": []}
        semanas[key]["total"] += float(f.get("monto_total") or 0)
        semanas[key]["facturas"].append(f.get("proveedor_nombre", ""))

    # Calcular promedio para niveles de alerta
    totales = [s["total"] for s in semanas.values()]
    promedio = sum(totales) / len(totales) if totales else 0

    result = []
    for semana in sorted(semanas.values(), key=lambda x: x["semana"]):
        nivel = "verde"
        if promedio > 0:
            if semana["total"] > promedio * 1.5:
                nivel = "rojo"
            elif semana["total"] > promedio * 1.1:
                nivel = "amarillo"
        semana["nivel"] = nivel
        result.append(semana)

    return result


@router.get("/proveedores-historico")
async def proveedores_historico(user_id: str):
    from app.services.supabase import get_supabase
    sb = get_supabase()

    provs = sb.table("proveedores").select("*").eq("user_id", user_id).order("score", desc=True).execute()

    result = []
    for p in provs.data:
        # Calcular % cumplimiento desde ratings
        ratings = sb.table("supplier_ratings").select("precio_cumplido, plazo_cumplido").eq("proveedor_id", p["id"]).execute()
        precio_ok = plazo_ok = total_r = 0
        for r in ratings.data:
            total_r += 1
            if r.get("precio_cumplido") is True:
                precio_ok += 1
            if r.get("plazo_cumplido") is True:
                plazo_ok += 1

        result.append({
            **p,
            "pct_precio_cumplido": round(precio_ok / max(total_r, 1) * 100),
            "pct_plazo_cumplido": round(plazo_ok / max(total_r, 1) * 100),
            "tasa_respuesta": round((p.get("total_respuestas") or 0) / max(p.get("total_solicitudes") or 1, 1) * 100),
        })

    return result


@router.get("/exportar-excel")
async def exportar_excel(user_id: str):
    """Genera Excel con estadísticas completas."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=503, detail="openpyxl no instalado. Ejecuta: pip install openpyxl")

    from app.services.supabase import get_supabase
    sb = get_supabase()

    wb = openpyxl.Workbook()

    # ── Hoja 1: OCs ───────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ordenes de Compra"
    headers1 = ["N° OC", "Proveedor", "Item", "Cantidad", "Precio Unit.", "Total", "Moneda", "Estado", "Fecha"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    ocs = sb.table("ordenes_compra").select("*").eq("user_id", user_id).order("created_at", desc=True).execute()
    for row_i, oc in enumerate(ocs.data, 2):
        ws1.append([
            oc.get("numero_oc"), oc.get("proveedor_nombre"), oc.get("nombre_item"),
            oc.get("cantidad"), oc.get("precio_unitario"), oc.get("precio_total"),
            oc.get("moneda"), oc.get("estado"), str(oc.get("created_at", ""))[:10],
        ])

    # ── Hoja 2: Proveedores ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Proveedores")
    headers2 = ["Proveedor", "Email", "Score", "Categoría", "Solicitudes", "OCs Enviadas", "OCs Confirmadas", "Tasa Resp. %"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h).font = Font(bold=True)

    provs = sb.table("proveedores").select("*").eq("user_id", user_id).order("score", desc=True).execute()
    for p in provs.data:
        tasa = round((p.get("total_respuestas") or 0) / max(p.get("total_solicitudes") or 1, 1) * 100)
        ws2.append([p.get("nombre"), p.get("email"), p.get("score"), p.get("categoria_score"), p.get("total_solicitudes"), p.get("total_oc_enviadas"), p.get("total_oc_confirmadas"), tasa])

    # ── Hoja 3: Facturas ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Facturas")
    headers3 = ["Proveedor", "N° Factura", "Fecha", "Vencimiento", "Monto Neto", "IVA", "Total", "Moneda", "Estado"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h).font = Font(bold=True)

    facts = sb.table("facturas").select("*").eq("user_id", user_id).order("fecha_factura", desc=True).execute()
    for f in facts.data:
        ws3.append([f.get("proveedor_nombre"), f.get("numero_factura"), str(f.get("fecha_factura") or ""), str(f.get("fecha_vencimiento") or ""), f.get("monto_neto"), f.get("iva"), f.get("monto_total"), f.get("moneda"), f.get("estado")])

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=estadisticas_claria.xlsx"},
    )
